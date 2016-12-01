import xml.etree.ElementTree as et
import ifcopenshell
import openpyxl
import time
import datetime
from collections import defaultdict
import sqlite3 as lite
import sys
from os.path import isfile
import externalfunctions
import os

requirements = openpyxl.load_workbook('quality_requirements.xlsx')
schedule = et.parse('UITVOERINGSTIJDSCHEMA.xml') 
file = ifcopenshell.open('Situatie_totaal4D.ifc') 
root = schedule.getroot()
link_between_obj_task = file.by_type("ifcrelassignstoprocess")
progress = []
completed = []
related_object  = []	
items = defaultdict(list) # what does this do??!!
# Get the complete list of activities from the project schedule: name of each activity, percentageComplete and it's start date (Tuples)
for named_tasks in root.iter('{http://schemas.microsoft.com/project}Task'):
	value = named_tasks.findtext('{http://schemas.microsoft.com/project}PercentComplete')
	name = named_tasks.findtext('{http://schemas.microsoft.com/project}Name')
	summary = named_tasks.findtext('{http://schemas.microsoft.com/project}Summary')
	start_string = ""
	startxml = named_tasks.findtext('{http://schemas.microsoft.com/project}Start')
	for st in startxml: # convert the start date of activities into a string
		if st is not 'T':
			start_string = start_string + st
		else:
			break
	x = int(value)
	s = int(summary)
	if name is not None and s == 0 and x > 0 and x < 100:
		activity_set1 = (name,x,start_string)
		progress.append(activity_set1)
	if name is not None and s == 0 and x == 100:
		activity_set2 = (name,x,start_string)
		completed.append(activity_set2)	
# Create a dictionary with the acitivities and the associated ifc element from the 4D model
activity_object_related_other = defaultdict(list)
activity_object_related_wall = defaultdict(list)	
for check in link_between_obj_task:
	for ele in check.RelatedObjects:
		if ele.is_a("IfcWindow") or ele.is_a("IfcDoor"): # only select objects of concern
			activity_object_related_other[check.RelatingProcess.Name].append(ele.GlobalId)
		elif ele.is_a("Ifcwall"):
			activity_object_related_wall[check.RelatingProcess.Name].append(ele.GlobalId)
# Determine the acitivity, or comibination of, that require a check to be perfromed	(Hard input for the time being)	
Reading = requirements.get_sheet_by_name('mapping')
max_rows = Reading.max_row
max_columns = Reading.max_column
activity_has_check = defaultdict(list)
for ro in range(2,max_rows+1):
	for co in range(2,max_columns+1):
		if Reading.cell(row = ro, column = co).value is not None:
			activity_has_check[Reading.cell(row = ro, column = 1).value].append(Reading.cell(row = ro,column = co).value)
		else:
			break
# get all the window & door elements in an ifc file and create a database table called objects with the results
element_Window = file.by_type("IfcWindow")
element_Door = file.by_type("IfcDoor")
element_Wall = file.by_type("IfcWall")
storey = file.by_type("IfcRelContainedInSpatialStructure")
listfortaksinprogressother = ["Stelwerk begane grond","Stelwerk 1e verdieping","Vooropname intern"]
listfortaskscompletedother = ["Stelwerk begane grond","Stelwerk 1e verdieping","Opleveren"]
listfortaskscompletedwall = ["Lijmwerk Kalkzandsteen elementen","Lijmwerk 1e verdieping kalkzandsteen","Lijmwerk 2e verdieping kalkzandsteen"]
wallswithopenings = externalfunctions.findwallopening()		
if not isfile('Newexample.db'):
	con = lite.connect('Newexample.db')
	cur = con.cursor()
	cur.execute("CREATE TABLE objects (ID INTEGER PRIMARY KEY AUTOINCREMENT,GUID TEXT,ELEMENT TEXT,LOCATION TEXT,TYPE TEXT)")
	for window in element_Window:
		id = window.GlobalId
		win = "window"
		input1 = [0,0,0,0]
		input1[0] = id
		input1[1] = win
		input1[3] = window.Name
		for x in storey:
			for y in x.RelatedElements:
				if y.is_a("IfcWindow"):
					if y.GlobalId == id:
						val2 = x.RelatingStructure.Name
						input1[2] = val2										
		cur.execute("INSERT INTO objects VALUES (NULL,?,?,?,?)",input1)		
	for door in element_Door:
		id = door.GlobalId
		do = "door"
		input2 = [0,0,0,0]
		input2[0] = id
		input2[1] = do
		input2[3] = door.Name				
		for x in storey:
			for y in x.RelatedElements:
				if y.is_a("IfcDoor"):
					if y.GlobalId == id:
						val2 = x.RelatingStructure.Name
						input2[2] = val2
		cur.execute("INSERT INTO objects VALUES (NULL,?,?,?,?)",input2)					
	for wall in element_Wall:
		id = wall.GlobalId
		do = "wall"
		input3 = [0,0,0,0]
		input3[0] = id
		input3[1] = do
		input3[3] = wall.Name				
		input3[2] = "Storey-1"					
		cur.execute("INSERT INTO objects VALUES (NULL,?,?,?,?)",input3)			
# Create the checklists table, only needed once because the list is assumed to be not as flexible as the objects table
	checklist = [('Werkbezoek aan fabriek, eventueel meerdere bezoeken',),('Controle aanvoer kozijnen',),('Hebben de sparingen de juiste afmetingen (inclusief tolerentie)?',),('Zitten de sparingen op de juiste plaats?',),('Kozijnen en beglazing volledig afplakken',),('Kwaliteit (let op beschadigingen) bij plaatsen',),('Kwaliteit (let op beschadigingen) bij oplevering',),('Zijn attesten aanwezig',),('Binnenkozijnen schoongemaakt',)]
	cur.execute("CREATE TABLE requirements (ID INTEGER PRIMARY KEY AUTOINCREMENT,Check_required TEXT)")
	cur.executemany("INSERT INTO requirements VALUES (NULL,?)", checklist)
	cur.execute("CREATE TABLE results (ID INTEGER PRIMARY KEY AUTOINCREMENT,InspectionID INTEGER REFERENCES inspections(ID),reviewer TEXT,result TEXT,Comments TEXT,DATE_of_Inspection TEXT)")
	cur.execute("CREATE TABLE inspections (ID INTEGER PRIMARY KEY AUTOINCREMENT,ObjectID  REFERENCES objects(ID),CheckID INTEGER REFERENCES requirements(ID),Inspection_Date TEXT)")
# Create the inspection checks (instances) table
	for task in progress:
		for key in activity_object_related_other:	
# If task is in progress and it relates to an object from the ifc model	
			if task[0] == key:
# Check if mapped key (process) precondition is met: (This is for the doors and windows)
				for index in activity_has_check:
					if index in task[0]:
						for one in listfortaksinprogressother:
							if index == one:
								allchecks = activity_has_check[index]
								allobjects = activity_object_related_other[task[0]]
								for object in allobjects:
									for checks in allchecks:
										if checks == "Controle aanvoer kozijnen":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Kozijnen en beglazing volledig afplakken":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Kwaliteit (let op beschadigingen) bij oplevering":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Binnenkozijnen schoongemaakt":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
	for task in completed:
		for key in activity_object_related_other:
			if task[0] == key:
				for index in activity_has_check:
					if index in task[0]:
						for one in listfortaskscompletedother:
							if index == one:
								allchecks1 = activity_has_check[index]
								allobjects1 = activity_object_related_other[task[0]]
								for object in allobjects1:
									for checks in allchecks1:
										if checks == "Kwaliteit (let op beschadigingen) bij plaatsen":
											x = object
											y = checks
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Zijn attesten aanwezig":
											x = object
											y = checks
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)																			
# and this is for the walls										
	for task in completed:
		for key in activity_object_related_wall:
			if task[0] == key:
				for index in activity_has_check:
					if index in task[0]:
						allchecks2 = activity_has_check[index]
						allobjects2 = activity_object_related_wall[task[0]]
						for object in allobjects2:
							if object in wallswithopenings: # filter out walls without openings
								for checks in allchecks2:
									x = object
									y = checks
									task2 = (task[2],)
									cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
									first_id = cur.fetchone()
									if first_id is None:
										continue
									cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
									second_id = cur.fetchone()
									if second_id is None:
										continue
									total = first_id + second_id + task2
									cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)							
	con.commit()	
# Scenario if table of objects in already created, being careful for duplicates	
else:
	con = lite.connect('Newexample.db')
	cur = con.cursor()
	cur.execute("SELECT GUID FROM objects")
	obj = cur.fetchall()
	obje = []
	inspections = []
	storey = file.by_type("IfcRelContainedInSpatialStructure")
	for x in obj:
		obje.append(x[0])
	cur.execute("SELECT ObjectID,CheckID,Inspection_Date FROM inspections")
	instance_inspection = cur.fetchall()
	for y in instance_inspection:
		inspections.append(y)		
	for window in element_Window:
		if window.GlobalId not in obje:
			id = window.GlobalId
			win ="window"
			input4 = [0,0,0,0]
			input4[0] = id
			input4[1] = win
			input4[3] = window.Name
			for x in storey:
				for y in x.RelatedElements:
					if y.is_a("IfcWindow"):
						if y.GlobalId == id:
							val2 = x.RelatingStructure.Name
							input4[2] = val2		
			cur.execute("INSERT INTO objects VALUES (NULL,?,?,?,?)",input3)	
	for door in element_Door:
		if door.GlobalId not in obje:
			id = door.GlobalId
			do ="door"
			input5 = [0,0,0,0]
			input5[0] = id
			input5[1] = do
			input5[3] = door.Name				
			for x in storey:
				for y in x.RelatedElements:
					if y.is_a("IfcDoor"):
						if y.GlobalId == id:
							val2 = x.RelatingStructure.Name
							input5[2] = val2		
			cur.execute("INSERT INTO objects VALUES (NULL,?,?,?,?)",input4)	
	for wall in element_Wall:
		if wall.GlobalId not in obje:
			id = wall.GlobalId
			do = "wall"
			input6 = [0,0,0,0]
			input6[0] = id
			input6[1] = do
			input6[3] = wall.Name				
			input6[2] = "Storey-1"					
			cur.execute("INSERT INTO objects VALUES (NULL,?,?,?,?)",input3)			
# Create the inspection checks (instances) table, if any new are presemt (cautious for duplicate instances)			
	for task in progress:
		for key in activity_object_related_other:	
# If task is in progress and it relates to an object from the ifc model	
			if task[0] == key:
# Check if mapped key (process) precondition is met: (This is for the doors and windows)
				for index in activity_has_check:
					if index in task[0]:
						for one in listfortaksinprogressother:
							if index == one:
								allchecks = activity_has_check[index]
								allobjects = activity_object_related_other[task[0]]
								for object in allobjects:
									for checks in allchecks:
										if checks == "Controle aanvoer kozijnen":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											if total not in inspections: # make sure instance not already created
												cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Kozijnen en beglazing volledig afplakken":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											if total not in inspections: # make sure instance not already created
												cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Kwaliteit (let op beschadigingen) bij oplevering":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											if total not in inspections: # make sure instance not already created
												cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Binnenkozijnen schoongemaakt":
											x = object
											y = checks 
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											if total not in inspections: # make sure instance not already created
												cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
	for task in completed:
		for key in activity_object_related_other:
			if task[0] == key:
				for index in activity_has_check:
					if index in task[0]:
						for one in listfortaskscompletedother:
							if index == one:
								allchecks1 = activity_has_check[index]
								allobjects1 = activity_object_related_other[task[0]]
								for object in allobjects1:
									for checks in allchecks1:
										if checks == "Kwaliteit (let op beschadigingen) bij plaatsen":
											x = object
											y = checks
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											if total not in inspections: # make sure instance not already created
												cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)
										elif checks == "Zijn attesten aanwezig":
											x = object
											y = checks
											cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
											first_id = cur.fetchone()
											if first_id is None:
												continue
											cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
											second_id = cur.fetchone()
											if second_id is None:
												continue
											task2 = (task[2],)
											total = first_id + second_id + task2
											if total not in inspections: # make sure instance not already created
												cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)																			
# and this is for the walls										
	for task in completed:
		for key in activity_object_related_wall:
			if task[0] == key:
				for index in activity_has_check:
					if index in task[0]:
						allchecks2 = activity_has_check[index]
						allobjects2 = activity_object_related_wall[task[0]]
						for object in allobjects2:
							if object in wallswithopenings: # filter out walls without openings
								for checks in allchecks2:
									x = object
									y = checks
									task2 = (task[2],)
									cur.execute("SELECT ID FROM objects WHERE GUID = ?",(x,))
									first_id = cur.fetchone()
									if first_id is None:
										continue
									cur.execute("SELECT ID FROM requirements WHERE Check_required = ?",(y,))
									second_id = cur.fetchone()
									if second_id is None:
										continue
									total = first_id + second_id + task2
									if total not in inspections: # make sure instance not already created
										cur.execute("INSERT INTO inspections VALUES (NULL,?,?,?)",total)							
	con.commit()
if isfile(r"C:\Users\esper\Desktop\Tue\7CC10\Thesis\Thesis draft\flask\static\models\Situatie_totaal4D.dae"):
	pass
else:
	externalfunctions.make_collada()